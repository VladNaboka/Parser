import datetime

import dateparser
import openpyxl
import requests
from openpyxl import load_workbook
from bs4 import BeautifulSoup as BS


resF = openpyxl.reader.excel.load_workbook(filename="Resouce.xlsx")
resF.active = 0
sheet = resF.active
idF = 0

file = "Items.xlsx"
wb = load_workbook(file)
wb.active = 0
FF = wb.active

FF['A1'] = 'id'
FF['B1'] = 'res_id'
FF['C1'] = 'link'
FF['D1'] = 'title'
FF['E1'] = 'content'
FF['F1'] = 'nd_date'
FF['G1'] = 's_date'
FF['H1'] = 'not_date'

while(FF.max_row > 1):
        FF.delete_rows(2)

tags = ['div', 'h2', 'span', 'ul', 'h3', 'a', 'time']
article = ''

def TContent(article):
    for o in range(len(tags)):
        while True:
            objTag = el.find_all(f'{tags[o]}', {frozenset(article)})
            if objTag == []:
                print('Неверно - ', tags[o], article)
                break
            else:
                print('Верно - ', tags[o], article)
                return objTag

for i in range(2, len(sheet['B']) + 1):
    idR = str(sheet['A' + str(i)].value)
    nameR = str(sheet['B' + str(i)].value)
    urlR = str(sheet['C' + str(i)].value)
    tagR = str(sheet['D' + str(i)].value)
    bottomR = str(sheet['E' + str(i)].value)
    titleR = str(sheet['F' + str(i)].value)
    dateR = str(sheet['G' + str(i)].value)
    url = requests.get(f"{urlR}")
    soup = BS(url.content, "html.parser")
    for el in soup.select(f"{tagR}"):
        try:
            # dateContent = el.find_all(f'{tags[o]}', {f"{dateR}"})
            # textContent = el.find_all(f'{tags[o]}', {f'{bottomR}'})
            # titleContent = el.find_all(f'{tags[o]}', {f'{titleR}'})
            dateContent = TContent(article={f"{dateR}"})
            textContent = TContent(article={f'{bottomR}'})
            titleContent = TContent(article={f'{titleR}'})
            internalL = [
                 a.get('href') for a in el.find_all('a')
                 if a.get('href') and a.get('href')]
            print(f"{nameR}")
            for i in range(2, len(dateContent)):
                dateP = dateparser.parse(dateContent[i-2].text.strip(), date_formats=['%Y %B %d'], settings={'TIMEZONE': 'UTC'})
                idF += 1
                FF.append([f'{idF}', f'{idR}', f'{internalL[i-2]}', f'{titleContent[i-2].text}', f'{textContent[i-2].text.strip()}',
                            f'{int(dateP.timestamp())}', str(int(datetime.datetime.now().timestamp())), f'{dateP}'])
        except Exception as er:
            pass
wb.save(file)
wb.close()
resF.close()

